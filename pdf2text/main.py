from openpyxl import Workbook
from utils import *
import argparse

def write_excel(img_dict, excel_path = './result.xlsx'):

    wb = Workbook()
    ws = wb.active

    for key, val in img_dict.items():
        
        try:
            ws.merge_cells(start_row=key[0], end_row=key[1], start_column=key[2], end_column=key[3])
            ws.cell(key[0], key[2], img2text(val))
        except AttributeError:
            print(f'merge conflict at {key}')
            continue

    wb.save(excel_path)


if __name__ == "__main__":

    parser = argparse.ArgumentParser()
    parser.add_argument("--pdfdir", type=str, required= True, help='path to the PDF file you want to convert')
    parser.add_argument("--exceldir", type=str, default=None, help='path to save the excel output, save to current directory by default')
    parser.add_argument('--lang', type=str,default='chi_sim+deu', help='Language(s) used in your PDF,the order indicates priority')
    parser.add_argument('--page', type=int, default=None, help='which page do you want to convert? Convert all pages by default, but could take quite some time~')
    parser.add_argument('--scale', type=int, default=50, help='See utils.py for more info')

    args = parser.parse_args()
    
    imgs = pdf2imgs(args.pdfdir)
    splitter = TableSplitter(args.scale)

    if args.page == None:
        i = 1
        for img in imgs:
            dict = splitter.split_image(img)
            write_excel(dict, f'./page-{i}.xlsx')
    else:
        dict = splitter.split_image(imgs[args.page - 1])
        write_excel(dict, f'./page-{args.page - 1}.xlsx')
    
    